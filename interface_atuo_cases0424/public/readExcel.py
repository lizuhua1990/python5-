# -*- coding: utf-8 -*-
# @Time    : 2018/4/26 13:03
# @Author  :
# @Email   :
# @File    : readExcel.py
# @Software: PyCharm


import os
from openpyxl import load_workbook

class readExcel:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet

    def read_Excel(self):
        rd=load_workbook(self.file)
        sheet=rd.get_sheet_by_name(self.sheet)

        #获取头部内容
        header=[]
        for i in range(sheet.max_column):
            header.append(sheet.cell(row=1,column=i+1).value)
        #return header

        data_list=[]
        for i in range(sheet.max_row-1):
            dict={}
            for ii in range(len(header)):
                dict[header[ii]]=sheet.cell(row=i+2,column=ii+1).value
            data_list.append(dict)
        return data_list


if __name__ == '__main__':
    print(readExcel(os.path.dirname(os.getcwd())+'/test_data/'+'testcases.xlsx','test_cases').read_Excel())