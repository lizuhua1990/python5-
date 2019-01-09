# -*- coding: utf-8 -*-
# @Time    : 2018/4/26 15:20
# @Author  :
# @Email   :
# @File    : writeExcel.py
# @Software: PyCharm

import os
from openpyxl import Workbook,load_workbook


class writeExcel:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet

    def write_Excel(self,i,ii,data):
        if os.path.exists(self.file):
            print('文件存在，不需要创建！')
            we=load_workbook(self.file)
            if self.sheet in we.get_sheet_names():
                print('表单存在，不需要创建！')
            else:
                print('表单不存在，已重新创建！')
                we.create_sheet(self.sheet)
                we.save(self.file)
        else:
            print('文件不存在，已重新创建！')
            we=Workbook()
            we.create_sheet(self.sheet)
            we.save(self.sheet)

        we_new=load_workbook(self.file)
        sheet=we_new.get_sheet_by_name(self.sheet)

        #先写头部
        #for i in range(len(header)):
            #sheet.cell(row=1,column=i+1).value=header[i]

        #写数据
        #for i in range(len(data)):
            #for ii in range(len(data[i])):
                #sheet.cell(row=i+2,column=ii+1).value=data[i][ii]
        sheet.cell (row=i, column=ii).value = data
        we_new.save(self.file)





if __name__ == '__main__':
    header = ['学号', '姓名', '性别', '班级']
    data = [['001', 'sunny', 'F', 'python5'], ['002', '小胖雨', 'M', 'python5']]
    print(writeExcel(os.path.dirname(os.getcwd())+'/test_data/'+'testcases.xlsx','test_cases1').write_Exevl(5,6,'hah'))