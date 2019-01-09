#写一个excel类，然后实例化对象，完成以下事项：
#1、判断excel文件是否存在，如若不存在，则创建一个excel文件，并创建sheet.
#2、在添加的sheet中添加python5班的各自小组的学员信息：学号、姓名、性别、班级。并保存数据
#3、将2中的数据全部读取出来，并且一行数据为一个字典，将所有数据都添加到列表中。
import os
from openpyxl import load_workbook
from openpyxl import Workbook
class excel_rw:
    def __init__(self,path,sheetname):#初始化参路径和表单名
        self.path=path
        self.sheetname=sheetname
    def pdExcle(self): #判断是否存在
        #不存在新建excl和sheet
        if os.path.splitext(self.path)[1]!=".xlsx" or os.path.exists(self.path)==False:
            workbook=Workbook()
            workbook.create_sheet(self.sheetname)
            workbook.save(self.path)
            print("新建表和表单成功")
        else:#否者就仅建一个sheet
            workbook1=load_workbook(self.path)
            if self.sheetname not in workbook1.get_sheet_names():
                workbook1.create_sheet(self.sheetname)
                workbook1.save(self.path)
                print("表存在，新建表单成功")
           # else:
                #print("表和表单都存在")

    def witeInfo(self,i,j,listinfo):
        #listinfo为学生数据list如[["学号","姓名","性别","班级"],["1","大白","男","python5"],["2","喵酱","女","python5"]]
        #第一个list为标题，后面为数据信息
        self.pdExcle()
        workbook2=load_workbook(self.path)
        sheet=workbook2.get_sheet_by_name(self.sheetname)
       # for i in range(1,len(listinfo)+1):#循环行数
            #for j in range(1,len(listinfo[0])+1):#循环列数
               # sheet.cell(i, j,listinfo[i-1][(j-1)])#写入信息
        sheet.cell(row=i,column=j).value=listinfo
        workbook2.save(self.path)
        #print("写入数据成功")
    def readInfo(self): #读取数据
        self.pdExcle()
        workbook2 = load_workbook(self.path)
        sheet = workbook2.get_sheet_by_name(self.sheetname)
        row=sheet.max_row   #确认行数
        column=sheet.max_column #确认列数
        list1=[]
        for i in range(2,row+1): #从第二行循环行数
            dict1 = {}
            for j in range(1,column+1):#循环列数
                result=sheet.cell(i, j).value #读取对应的字段值
                dict1[sheet.cell(1,j).value]=result #对应的key和字段值加入字典
            list1.append(dict1)#字典加入list
        #print(list1)
        return list1
if __name__=="__main__":
    path=os.path.join(os.getcwd(),"work.xlsx")
    excel_1=excel_rw(path,"work1")
    #excel_1.pdExcle()
    listinfo=[["学号","姓名","性别","班级"],["1","大白","男","python5"],["2","喵酱","女","python5"],["3","杰东","男","python5"]]
    excel_1.witeInfo(listinfo)
    print(excel_1.readInfo())

