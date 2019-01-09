__author__ = '20489'
'''写一个excel类，然后实例化对象，完成以下事项：
1、判断excel文件是否存在，如若不存在，则创建一个excel文件，并创建sheet.
2、在添加的sheet中添加python5班的各自小组的学员信息：学号、姓名、性别、班级。并保存数据
3、将2中的数据全部读取出来，并且一行数据为一个字典，将所有数据都添加到列表中。
注意：写成类！！！！类里面有各个功能的函数！！！'''
import os
from openpyxl import load_workbook
from interface_atuo_cases_teacher.public.read_config import readConfig
from interface_atuo_cases_teacher.conf import project_path

class R_W_Excel:
    def __init__(self,file,sheet_data,sheet_init):
        self.file=file
        self.sheet_data=sheet_data
        self.sheet_init=sheet_init
        self.ip=readConfig().get_value(project_path.http_conf_path,'HTTP','ip')

    def write_excel(self,i,result_dcit):#写数据到Excel里面
      #写入数据的操作
        wb_new=load_workbook(self.file)
        sheet=wb_new.get_sheet_by_name(self.sheet_data)
        sheet.cell(i,7).value=result_dcit['Actual_code']
        sheet.cell(i,8).value=result_dcit['sql_result']
        sheet.cell(i,9).value=result_dcit['Result']
        sheet.cell(i,10).value=result_dcit['Reason']
        wb_new.save(self.file)

    #读取初始化手机号
    def get_init_phone(self):
        wb=load_workbook(self.file)
        sheet=wb[self.sheet_init]
        initphone=sheet.cell(row=1,column=2).value
        return initphone

     #3、将2中的数据全部读取出来，并且一行数据为一个字典，将所有数据都添加到列表中。
    def read_excel(self,mode,case_list):
        wb=load_workbook(self.file)
        sheet=wb.get_sheet_by_name(self.sheet_data)

        initphone=self.get_init_phone()
        secondphone=str(int(initphone)+1)

        data_list=[]#存储测试数据
        if mode=='1':
            #获取数据并且返回数据
            for i in range(2,sheet.max_row+1):
                dict={}
                dict['id']=sheet.cell(row=i,column=1).value
                dict['method']=sheet.cell(row=i,column=2).value
                dict['url']=self.ip+sheet.cell(row=i,column=3).value
                dict['data']=eval(sheet.cell(row=i,column=4).value)
                dict['sql']=eval(sheet.cell(row=i,column=5).value)
                #加一步操作，考虑字典里面mobilephone这个key是否存在的情况
                if dict['data']['mobilephone']=='initphone':
                    dict['data']['mobilephone']=initphone
                    dict['sql']['condition']=initphone
                elif dict['data']['mobilephone']=='secondphone':
                    dict['data']['mobilephone']=secondphone
                    dict['sql']['condition']=secondphone
                else:
                    print("不需要做任何操作，因为没有变更手机号或者是不需要手机号")
                dict['Expected_code']=sheet.cell(row=i,column=6).value
                data_list.append(dict)
            self.update_phone(str(int(initphone)+2))

        elif mode=='0':
            #id是用例编号  case_list里面都是用例的编号
            for id in case_list:#确保你的case_list是列表类型的数据 2
                dict={}
                dict['id']=sheet.cell(row=id+1,column=1).value
                dict['method']=sheet.cell(row=id+1,column=2).value
                dict['url']=self.ip+sheet.cell(row=id+1,column=3).value
                dict['data']=sheet.cell(row=id+1,column=4).value
                dict['sql']=eval(sheet.cell(row=id+1,column=5).value)
                #加一步操作，考虑字典里面mobilephone这个key是否存在的情况
                if dict['data']['mobilephone']=='initphone':
                    dict['data']['mobilephone']=initphone
                    dict['sql']['condition']=initphone
                elif dict['data']['mobilephone']=='secondphone':
                    dict['data']['mobilephone']=secondphone
                    dict['sql']['condition']=secondphone
                else:
                    print("不需要做任何操作，因为没有变更手机号或者是不需要手机号")
                dict['Expected_code']=sheet.cell(row=id+1,column=6).value
                data_list.append(dict)
            self.update_phone(str(int(initphone)+2))
        return data_list

    def update_phone(self,phone):
         wb=load_workbook(self.file)
         sheet=wb.get_sheet_by_name(self.sheet_init)
         sheet.cell(1,2).value=phone
         wb.save(self.file)

if __name__=="__main__":
    header=['学号','姓名','性别','班级']
    data=[['001','sunny','F','python5'],['002','小胖鱼','M','python4']]
    R_W_Excel('python5.xlsx','python5').write_excel(header,data)
    print(R_W_Excel('python5.xlsx','python5').read_excel())









