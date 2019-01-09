# -*- coding: utf-8 -*-
# @Time    : 2018/4/26 13:03
# @Author  :
# @Email   :
# @File    : readExcel.py
# @Software: PyCharm


import os
from interface_atuo_cases0505.public.config import config
from openpyxl import load_workbook

mode = config ().read_config (os.path.dirname (os.getcwd ()) + '/conf/' + 'case.conf', 'FLAG', 'mode')
#print(mode)

class readExcel:
    def __init__(self,file,sheet,init):
        self.file=file
        self.sheet=sheet
        self.init=init

    def read_Excel(self):
        rd=load_workbook(self.file)
        sheet=rd.get_sheet_by_name(self.sheet)

        initphone=self.get_init_phone()
        print(initphone)#当作一个局部变量来看这个，一轮循环下来后在更新到excel

        data_list=[]
        for i in range(sheet.max_row-1):
            data=eval(sheet.cell(row=i+2,column=2).value)#取值excel中的字典
            code=sheet.cell(row=i+2,column=3).value
            if data['mobilephone']==initphone:#判断是否与保存的最后一次相同
                data['mobilephone'] = initphone+1#相同就加一
            else:
                data['mobilephone'] = initphone#给字典赋值变量中的号码
            initphone += 1#作为局部变量累加存储
            dict={'id':i+1,'data':data,'code':code}
            data_list.append(dict)
        #self.update_phone(initphone)#把最后的值更新到excel中
        print(data_list)
        return data_list

    #读取初始化手机
    def get_init_phone(self):
        wb=load_workbook(self.file)
        sheet=wb[self.init]
        initphone=sheet.cell(row=1,column=2).value
        return initphone
    def update_phone(self,phone):#更新手机
        wb = load_workbook (self.file)
        sheet=wb.get_sheet_by_name(self.init)
        sheet.cell(1,2).value=phone
        wb.save(self.file)


#if __name__ == '__main__':
#    print(readExcel(os.path.conf.dirname(os.getcwd())+'/test_data/'+'testcases.xlsx','test_cases','init').read_Excel())
