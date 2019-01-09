# -*- coding: utf-8 -*-
# @Time    : 2018/4/17 23:55
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : class_excel_write.py
# @Software: PyCharm


#操作情景是往一个新的Excel里写入数据
from openpyxl import Workbook#写入数据到Excel里面

#第一步：新建一个Excel
workbook=Workbook()#新建目标Excel文件名

#第二步：新建一个表单
sheet=workbook.create_sheet('python5_2')

#第三步：保存
workbook.save('python5_2.xlsx')

#写入数据
from openpyxl import load_workbook

workbook_2=load_workbook('python5_2.xlsx')
sheet=workbook.get_sheet_by_name('python5_2')
sheet.cell(row=1,column=2).value='333'#赋值到第一行第二列
workbook.save('python5_2.xlsx')

