# -*- coding: utf-8 -*-
# @Time    : 2018/4/18 8:57
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : class_excel_rw.py
# @Software: PyCharm


#引入的模块
from openpyxl import load_workbook#专门用来打开已经存在的Excel
from openpyxl import Workbook#专门新建的Excel

#对Excel的操作，总的来说有2种，一种是excel存在，一种是excel不存在要新建

#新建Excel的过程
workbook=Workbook()
workbook.create_sheet('python5')
workbook.save('python_3.xlsx')

#读写的操作  可以同时进行
#打开Excel
workbook_rw=load_workbook("python5.xlsx")
sheet_rw=workbook_rw.get_sheet_by_name('python5')

#读操作
read_result=sheet_rw.cell(row=4,column=2).value
print(read_result)

#写操作===保存
sheet_rw.cell(row=4,column=2).value='小小豆'
workbook_rw.save("prthon5.xlsx")

