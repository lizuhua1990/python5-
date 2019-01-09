# -*- coding: utf-8 -*-
# @Time    : 2018/4/18 11:26
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : homework0414.py
# @Software: PyCharm

#4-14 超级简单的作业~Excel读取数据类

#写一个excel类，然后实例化对象，完成以下事项：
#1、判断excel文件是否存在，如若不存在，则创建一个excel文件，并创建sheet.
#2、在添加的sheet中添加python5班的各自小组的学员信息：学号、姓名、性别、班级。并保存数据
#3、将2中的数据全部读取出来，并且一行数据为一个字典，将所有数据都添加到列表中。
#注意：写成类！！！！类里面有各个功能的函数！！！

"""
1.1.读取配置文件
-read(filename) 直接读取ini文件内容
-sections() 得到所有的section，并以列表的形式返回
-options(section) 得到该section的所有option
-items(section) 得到该section的所有键值对
-get(section,option) 得到section中option的值，返回为string类型
-getint(section,option) 得到section中option的值，返回为int类型

1.2.写入配置文件
-add_section(section) 添加一个新的section
-set( section, option, value) 对section中的option进行设置
  需要调用write将内容写入配置文件。

Python的ConfigParser Module中定义了3个类对INI文件进行操作。
分别是RawConfigParser、ConfigParser、SafeConfigParser。
RawCnfigParser是最基础的INI文件读取类，ConfigParser、SafeConfigParser支持对%(value)s变量的解析

"""
from openpyxl import *
from configparser import *

class homework0414:
    cla_config=ConfigParser()#创建读取配置文件实例
    #cla_rwconfig=RawConfigParser()#可以删除

    def __init__(self,cla_ConfigName):# 初始化配置文件名称
        self.cla_ConfigName=cla_ConfigName

    def sureConfig(self,cla_section,cla_option):
        self.cla_config.read(self.cla_ConfigName)#读取配置文件中内容
        return self.cla_config.get(cla_section,cla_option)#返回获取到的标签名  和 键值名key

    def sureExcel(self,cla_section,cla_option):
        # 读取conf配置文件中的xlsx文件名称和sheetName
        self.excelName=self.sureConfig(cla_section,cla_option['excelName'])
        self.sheetName=self.sureConfig(cla_section,cla_option['sheetName'])
        print(self.excelName,self.sheetName)
        # 判断xlsx文件是否存在，不存在则创建
        try:
            open(self.excelName,'r',encoding='utf-8').close()
        except:
            print('指定%s文件不存在'%self.excelName)
            open (self.excelName, 'w', encoding='utf-8').close ()


    def writeExcel(self,cla_data):
        workbook_1 = load_workbook(self.excelName)
        workbook_1.create_sheet(self.sheetName)
        sheet_1 = workbook_1.get_sheet_by_name(self.sheetName)
        #写入数据
        colNum = 0
        for dic_data in cla_data:
            rowNum = 1
            colNum += 1
            for i,ii in dic_data.items():
                sheet_1.cell(row=rowNum, column=colNum).value = i
                for data in ii:
                    rowNum += 1
                    sheet_1.cell(row=rowNum, column=colNum).value = data
            workbook_1.save(self.excelName)

    def readExcel(self):
        list_data = []
        # 加载xlsx文件
        cla_workbook = load_workbook(self.excelName)
        # 获取xlsx中的sheet
        cla_sheet = cla_workbook.get_sheet_by_name(self.sheetName )
        # 获取总行数
        int_row = cla_sheet.max_row
        # 获取总列数
        int_column = cla_sheet.max_column
        # 遍历所有单元格，组成字典放到list中
        for line in range(1,int_row):
            for i in range(1,int_column+1):
                str_key = cla_sheet.cell(row=1, column=i).value
                str_val = cla_sheet.cell(row=line+1, column=i).value
                list_data.append({str_key:str_val})
        return list_data


if __name__ == '__main__':
    hh = homework0414("cla_Student.conf")
    Name = hh.sureExcel('TESTCASEFILENAME',{'excelName':'excelName','sheetName':'sheetName'})
    #print(hh.sureExcel('TESTCASEFILENAME',{'excelName':'excelName','sheetName':'sheetName'}))
    hh.writeExcel([
        {'id':['0501','0506','0518','0519']},
        {'name':['喵酱','老宝儿','呼呼','时光']},
        {'性别':['女','女','男','男']},
        {'班级':['A','A','A','A']}])
    print(hh.readExcel())




















