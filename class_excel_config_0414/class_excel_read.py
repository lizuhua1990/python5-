# -*- coding: utf-8 -*-
# @Time    : 2018/4/17 23:01
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : class_excel_read.py
# @Software: PyCharm


#workbooke工作簿====sheet  表====cell单元格    行   列

from openpyxl import load_workbook#读取Excel 才需要的

#第一步：打开Excel
workbook=load_workbook('python5.xlsx')#你要打开的目标Excel文件名
#相对路径：当前路径下打开这个文件‘python5.xls’

#workbook=load_workbook('D:\python5\class_excel_RW0414\python5.xls')
#绝对路径


#第二步：确定你要读取的数据的表单（要定位一个表单）
sheet=workbook.get_sheet_by_name('python5')

#第三步：确认你要都的数据 定位坐标去读取
result=sheet.cell(row=1,column=1).value#赋值操作
#1、索引或者是下标 是从1开始的
#2、row  colum

#行列的获取
row=sheet.max_row#获取行数
col=sheet.max_column#获取列数
print(result,row,col)

for i in range(2,row+1):#这个模块是从1开始的
    print(sheet.cell(row=i,column=2).value)





























