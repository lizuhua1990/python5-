# -*- coding: utf-8 -*-
# @Time    : 2018/4/22 14:51
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : class_Excel.py
# @Software: PyCharm
'''
写一个excel类，然后实例化对象，完成以下事项：
1、判断excel文件是否存在，如若不存在，则创建一个excel文件，并创建sheet.
2、在添加的sheet中添加python5班的各自小组的学员信息：学号、姓名、性别、班级。并保存数据
3、将2中的数据全部读取出来，并且一行数据为一个字典，将所有数据都添加到列表中。
'''

import os
from openpyxl import Workbook
from openpyxl import load_workbook

class R_W_Excel:
    def __init__(self,file,sheet):
        self.sheet=sheet
        self.file=file

    def write_excel(self,header,data):
        #data必须是list类型列表  header必须是list类型列表
        if os.path.exists(self.file):#os返回值本身就是布尔值
            print('excel已经存在，不需要创建！')
            #判断表单是否存在
            #1、获取所有的表单
            wb=load_workbook(self.file)
            if self.sheet in wb.get_sheet_names():
                print('表单已经存在')
            else:
                wb.create_sheet(self.sheet)
                wb.save(self.file)

        else:
            print('excel不存在，需要重新创建！')
            #创建excel
            wb=Workbook()
            wb.create_sheet(self.sheet)
            wb.save(self.file)

        #写入数据的操作
        wb_new=load_workbook(self.file)
        sheet=wb_new.get_sheet_by_name(self.sheet)

        #先写头部
        for i in range(len(header)):
            sheet.cell(row=1,column=i+1).value=header[i]

        #写数据
        for i in range(len(data)):#0 1
            for ii in range(len(data[i])):
                sheet.cell(row=i+2,column=ii+1).value=data[i][ii]

        wb_new.save(self.file)

    #将数据全部读取出来，并且一行数据为一个字典，将所有数据都添加到列表中
    def read_excel(self):
        wb=load_workbook(self.file)
        sheet=wb.get_sheet_by_name(self.sheet)

        #获取头部
        header=[]
        for i in range(sheet.max_column):
            header.append(sheet.cell(row=1,column=i+1).value)

        #获取数据并返回
        data_list=[]
        for i in range(sheet.max_row-1):
            dict={}
            for ii in range(len(header)):
                dict[header[ii]]=sheet.cell(row=i+2,column=ii+1).value
            data_list.append(dict)
        return data_list


if __name__ == '__main__':
    header = ['学号', '姓名', '性别', '班级']
    data = [['001', 'sunny', 'F', 'python5'], ['002', '小胖雨', 'M', 'python5']]
    R_W_Excel('python5.xlsx','python5').write_excel(header,data)

    print(R_W_Excel('python5.xlsx','python5').read_excel())












